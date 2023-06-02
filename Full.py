#1. Install python and pip
    #1. Go to this link, scroll to the bottom and click on the proper installer for your system https://www.python.org/downloads/release/python-3108/
    #2. Open the installer and check add to PATH, then choose install.
    #3. To ensure nothing was corrupted in the download, type in the following commands in cmds
        #1. powershell
        #2. python --version
        #3. If the version came up then everything is well
    #4. Check to see if pip is installed by running the following commands in cmds
        #1. pip --version
        #2. If it outputted the verison of pip, it is already installed.
        #3. If the previous step did not happen to you, here is how you can install it.
            #1.  py -m ensurepip --upgrade
#2. Enable the API's needed by going to https://console.cloud.google.com/cloud-resource-manager
    #1. In the search bar type in "Google Drive API"
    #2. Click "Create Project". The Name of the project does not matter
    #3. After creating the project, click "Enable" on the google drive API
    #4. In the search bar type in "Google Sheets API"
    #5. Click "Enable"
#3. Create a service account by going to https://console.cloud.google.com/apis/credentials
    #1. After going to the page, on the top next to delete you should see "Create Credentials". Click on it and choose "Service Account"
    #2. Naming of it does not matter nor does the description. When you're done click "Create and Continue"
    #3. Click the blue button "Done"
    #4. Click the pencil under the actions column on the same row of the bot you just created.
    #5. Navigate to the "Keys" tab
    #6. Click the dropdown "Add Key", and click "Create New key" Choose the JSON file
    #7 Copy the email address shown under "Email" in the "Details" tab
    #8 Navigate to the inventory sheet and share the spreadsheet with the bot, make sure to make it as an editor.
#4. Download the needed libraries by running the following code in cmds
    #1. pip install gspread, openpyxl, selenium, webdriver-manager, Pillow
    #2. Reopen Visual Studio's to ensure that the download took effect (only if you're using visual studios)
#5. Insert the needed information inside of the quotes. Where to put the information will be noted with the corresponding step#
    #1. Path to the service account
    #2. Path needed for a text file (the text file does not have to be created, the code will do it for you). This will be the logs
    #3. Chromebook's spreadsheet title, this is the whole spreadsheet's title located in the top left corner.
    #4. Ipad's spreadsheet title, this is the whole spreadsheet's title located in the top left corner.
    #5. The Sold excel file path
    #6. The path of the excel sheet with the asset tags you want retired
#6. Run the script
    #1. Run the script after inserting the asset tags you want transfered inside of the target asset tags excel file
    #2. Check the logs to see if there's any missing asset tags
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from selenium.webdriver.common.keys import Keys
import gspread
import tkinter as tk
from tkinter import *
from tkinter import messagebox
from tkinter.scrolledtext import *
from PIL import ImageTk, Image
from datetime import datetime
from time import sleep

pathServiceAccount = r"path" #1
pathTxt = r"path" #2
chromeSpreadsheetName = "Copy of Copy of Inv Tracker Copy 2022-12-23 07:47:20" #3
ipadSpreadsheetName = 'Copy of Ipad Inventory' #4
sellSheetName = 'Current Year'
workbookSell = r"path" #5
workbookTargetAssets = r"path" #6
currentlyRunning = False #determines if it's currently running or not
togs = [] #Table for all of the active toggles
serviceAccount = gspread.service_account(filename = pathServiceAccount) #loads the service account
chromeGoogleSpreadsheet = serviceAccount.open(chromeSpreadsheetName) #Opens the chromebook inventory spreadsheet
targetWorkbook = load_workbook(workbookTargetAssets) #Loads excel worksheet file for 
targetWorksheet = targetWorkbook.active #Gets active worksheet to pull asset tags from
sellWorkbook = load_workbook(workbookSell) #Loads sell excel file
sellWorksheet = sellWorkbook[sellSheetName] #Active worksheet inside of the workbook
sheetName = 'Chromebooks'
forceRetire = False #Forcefully retires device, defaults to false
invSheetName = None
sheets = chromeGoogleSpreadsheet.worksheet('Chromebooks')
sheetsToLook = [chromeGoogleSpreadsheet.worksheet('Chromebooks')]
assetChromebooks = [] #Target assets of Chromebooks to find
assetIpads = [] #Target assets of Ipads to find
serialIpads = {} #Dictionary to keep track of serial numbers
rowsFound = [] #Table for the assets found and their information
tagsNotFound = [] #Table for assets not found
allInfo = None
driver = []
invSheetDd = None
username = None
password = None
animationCorrector = 0
percentDone = 0

def sonicRun():
    global animationCorrector
    if animationCorrector == 0:
        sonicLabel.configure(image=sonicRunFull[animationCorrector])
    elif animationCorrector < 8:
        sonicLabel.configure(image=sonicRunFull[animationCorrector])
    else:
        animationCorrector = 0
        sonicLabel.configure(image=sonicRunFull[animationCorrector])
    animationCorrector += 1

def getTargets():
    for index, cell in enumerate(targetWorksheet['A']): #Grabs all the values inside of the recieving worksheet
        if cell.value != None and cell.value != 'Chromebooks': #Checks for empty cells and ignores the Chromebooks cell
            assetChromebooks.insert(index, cell.value) #Inserts the values into "assetChromebooks"
    for index, cell in enumerate(targetWorksheet['B']): #Grabs all the values inside of the recieving worksheet
        if cell.value != None and cell.value != 'Ipads': #Checks for empty cells and ignores the Ipads cell
            assetIpads.insert(index, cell.value) #Inserts the values into "assetIpads"

def elementInteract(method, element, clickOrKeys, keysToSend = None):
    if clickOrKeys:
        driver.find_element(method, element).click()
    else:
        driver.find_element(method, element).send_keys(keysToSend)

def waitForExistance(elementXPath, waittime):
    try:
        waitFor = WebDriverWait(driver, waittime).until(EC.presence_of_element_located((By.XPATH, elementXPath)))
        return True
    except:
        return False

def waitUntilClickable(element, waittime):
    try:
        WebDriverWait(driver, waittime).until(EC.element_to_be_clickable((By.XPATH, element)))
        return True
    except:
        return False

def openSellWorkbook():
    global sellWorkbook
    global sellWorksheet
    sellWorkbook = load_workbook(workbookSell)
    sellWorksheet = sellWorkbook[sellSheetName]

def sheetstoexcel(sheetname): #full function for moving information from the inventory tracker to the sell file
    global assetChromebooks
    global rowsFound
    global tagsNotFound
    global serialIpads
    placeHolderTable = []
    corrector = 0
    if assetChromebooks:
        chromeInventorySheet = chromeGoogleSpreadsheet.worksheet(sheetname) #Sheet open inside of the main spreadsheet
        chromeAllInfo = chromeInventorySheet.get_all_records()
        for asset in assetChromebooks:
            placeHolderTable.append(asset)
        for i, rows in enumerate(chromeAllInfo):
            if rows['Asset'] in placeHolderTable:
                sleep(0.9)
                placeHolderTable.remove(rows['Asset'])
                sellWorksheet.append(list(rows.values())) #Enters the data into the next open row in the excel sheet
                sellWorkbook.save(workbookSell) #Saves the workbook
                log( "Found and Written " + str(rows['Asset']) + '\n') #Enters that it found the certain asset tag
                chromeInventorySheet.delete_rows(i + 2 - corrector)
                corrector += 1
                sonicRun()
        for leftOvers in placeHolderTable: #Every asset that wasn't found
            log('Did not find asset# '+str(leftOvers) + ' in inventory' + '\n') #Writes the asset into logs
            sonicRun()
        placeHolderTable = []
        corrector = 0
    else:
        log('No Chromebook targets')

def log(text):
    txtWrite = open(pathTxt, 'a')
    txtWrite.write(text + '\n') #Writes in logs
    textLogs.insert(END, text + '\n')
    txtWrite.close()
    root.update()
def removeFilter():
    elemRemoveFilter = driver.find_element('xpath', '//*[@id="mCnz0"]/div/div/div[2]/span/div/div[3]/div/div/div/div[2]/ul/li[2]/div')
    driver.execute_script('arguments[0].click();', elemRemoveFilter) #Removes the asset tag searched

def destinyFull():
    global forceRetire
    options = Options()

    options.add_experimental_option('detach', True) #closes window on completion
    options.add_argument('--incognito')
    options.headless = False #If selected true, it runs in the background

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()),
    options = options) #Chooses Chrome as the browser
    driver.implicitly_wait(5)

    driver.get("https://library.seymour.k12.wi.us/district/servlet/presentdistrictloginform.do") #Goes to the destiny login page
    driver.find_element('xpath', '//*[@id="Login"]').click() #Clicks to go to the login screen
    driver.find_element('xpath', '//*[@id="ID_loginName"]').send_keys('#######') #Enters username
    driver.find_element('xpath', '//*[@id="ID_password"]').send_keys('#######') #Enters password
    driver.find_element('xpath', '/html/body/table/tbody/tr[4]/td/table/tbody/tr/td/form/table/tbody/tr[3]/td/table/tbody/tr/td/table/tbody/tr[4]/td/input[1]').click() #Clicks login
    driver.find_element('xpath', '//*[@id="rightTableSites_0"]/tbody/tr/td/ul/li/span/a[1]').click() #Clicks SMS as the location
    driver.get('https://library.seymour.k12.wi.us/cataloging/servlet/presentadvancedsearchredirectorform.do?l2m=Library%20Search&tm=TopLevelCatalog') #Goes to the catalog search
    driver.find_element('xpath', '//*[@id="Resource Search"]').click() #Clicks on the resource search
    driver.find_element('xpath', '//*[@id="searchFieldsTable"]/tbody/tr/td[1]/select/optgroup[2]/option[1]').click() #Chooses the barcode option
    def checkExistance(xpath): #Checks to see if a element exists
        try:
            driver.find_element('xpath', xpath)
        except NoSuchElementException:
                return False
        return True
    for i, asset in enumerate(assetChromebooks): #Goes through this process with each one of the assets
        sonicRun()
        if len(str(asset)) == 4:
            driver.find_element('xpath', '/html/body/table[2]/tbody/tr[3]/td/table/tbody/tr/td[2]/table/tbody/tr/td[2]/table/tbody/tr/td/form/table[2]/tbody/tr/td/table/tbody/tr[2]/td/table/tbody/tr[2]/td[2]/table/tbody/tr/td[2]/input').send_keys('0' + str(asset)) #Types in the asset# + a 0 at the start
        elif len(str(asset)) > 4:
            driver.find_element('xpath', '/html/body/table[2]/tbody/tr[3]/td/table/tbody/tr/td[2]/table/tbody/tr/td[2]/table/tbody/tr/td/form/table[2]/tbody/tr/td/table/tbody/tr[2]/td/table/tbody/tr[2]/td[2]/table/tbody/tr/td[2]/input').send_keys(str(asset))
        driver.find_element('xpath', '//*[@id="buttonSearch"]').click() #Searches
        if checkExistance('//*[@id="messageBox"]/tbody/tr[1]/td[2]'): #If the chromebook isn't already at the SMS site
            if driver.find_element('xpath', '//*[@id="messageBox"]/tbody/tr[1]/td[2]').get_attribute('innerHTML').__contains__('Please note'): #Makes sure that it's asking the question if you wanted to move it
                driver.find_element('xpath', '//*[@id="divToHide"]/input[1]').click() #Moves it to the SMS site
        if checkExistance('//*[@id="messageBox"]/tbody/tr/td'): #Checks to see if the element saying the chromebook isn't inside of destiny exists
            log(str(asset) + ' is not avalaible') #Writes in logs saying the Chromebook wasn't avalaible in destiny
            sonicRun()
            driver.find_element('xpath', '//*[@id="Resource Search"]').click() #Clicks back to the resource search
            continue #Restarts the process with the next asset in line
        if driver.find_element('xpath', '//*[@id="itemTable_0"]/tbody/tr[2]/td[3]/a').get_attribute('innerHTML').__contains__('Available'): #Checks to see if the Chromebook isn't checked out or already retired
            driver.find_element('xpath', '//*[@id="itemTable_0"]/tbody/tr[2]/td[6]/a[2]/img').click() #clicks on the edit button
            log(str(asset) + ' Has been retired ' + ':' + ')' + '\n')
            sonicRun()
        elif driver.find_element('xpath', '//*[@id="itemTable_0"]/tbody/tr[2]/td[3]/a').get_attribute('innerHTML').__contains__('Checked') and forceRetire:
            driver.find_element('xpath', '//*[@id="TopLevelCirculation"]').click()
            driver.find_element('xpath', '//*[@id="Check In Items"]').click()
            if len(str(asset)) == 4:
                driver.find_element('xpath', '//*[@id="headerTable"]/tbody/tr/td[2]/input').send_keys('0' + str(asset))
            else:
                driver.find_element('xpath', '//*[@id="headerTable"]/tbody/tr/td[2]/input').send_keys(str(asset))
            driver.find_element('xpath', '//*[@id="go"]').click()
            if checkExistance('//*[@id="messageBox"]/tbody/tr/td[2]'):
                driver.find_element('xpath', '//*[@id="ComponentsTable"]/tbody/tr[5]/td/input[1]').click()
            log(str(asset) + ' was  ' + driver.find_element('xpath', '/html/body/table[2]/tbody/tr[3]/td/table/tbody/tr/td[2]/table/tbody/tr/td[2]/table/tbody/tr/td/form/div[3]/table/tbody/tr[3]/td[2]/span[1]').get_attribute('innerHTML') + ' but is now checked in '+ '\n')
            sonicRun()
            driver.find_element('xpath', '//*[@id="TopLevelCatalog"]').click()
            driver.find_element('xpath', '//*[@id="Resource Search"]').click()
            driver.find_element('xpath', '//*[@id="searchFieldsTable"]/tbody/tr/td[1]/select/optgroup[2]/option[1]').click()
            if len(str(asset)) == 4:
                print(len(str(asset)))
                driver.find_element('xpath', '//*[@id="searchFieldsTable"]/tbody/tr/td[2]/input').send_keys('0' + str(asset))
            else:
                driver.find_element('xpath', '//*[@id="searchFieldsTable"]/tbody/tr/td[2]/input').send_keys(str(asset))
            driver.find_element('xpath', '//*[@id="buttonSearch"]').click()
            driver.find_element('xpath', '//*[@id="itemTable_0"]/tbody/tr[2]/td[6]/a[2]/img').click()
        else: #If it isn't available
            log(str(asset) + ' is marked as ' + str(driver.find_element('xpath', '//*[@id="itemTable_0"]/tbody/tr[2]/td[3]/a').get_attribute('innerHTML')) + ' and needs to be done manually') #Writes in logs
            sonicRun()
            driver.find_element('xpath', '//*[@id="Resource Search"]').click() #Clicks back on reasource search
            continue
        driver.find_element('xpath', '//*[@id="tableMain"]/tbody/tr[4]/td[2]/table/tbody/tr/td[1]/select/option[9]').click() #Checks status to retired
        driver.find_element('xpath', '//*[@id="saveCopy"]').click() #Clicks save
        driver.find_element('xpath', '//*[@id="Resource Search"]').click() #Goes back to resource search
        driver.find_element('xpath', '//*[@id="searchFieldsTable"]/tbody/tr/td[1]/select/optgroup[2]/option[1]').click() #clicks on barcode option

def googleAdmin():
    def mainGoogleAdmin():
        global username
        global password
        global driver
        options = Options()
        options.add_experimental_option('detach', True) #Closes window on completion
        #options.add_argument('--headless') #If enabled, it runs in the background
        options.add_argument("--incognito") #Runs in incognito mode so no autofill usernames/passwords gets in the way
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()),
        options = options) #Chooses Chrome as the browser
        driver.implicitly_wait(10) #Waits 10 seconds to find an element befor sending an error
        driver.get("https://admin.google.com/ac/chrome/devices/")
        try:
            elementInteract('xpath', '/html/body/div[1]/div[1]/div[2]/div/c-wiz/div/div[2]/div/div[1]/div/form/span/section/div/div/div[1]/div/div[1]/div/div[1]/input', False, username) #Types username
            elementInteract('xpath', '/html/body/div[1]/div[1]/div[2]/div/c-wiz/div/div[2]/div/div[2]/div/div[1]/div/div/button/span', True) #Clicks next
            try: #Checks if the password box that isn't the school login portal exists
                elementInteract('xpath', '//*[@id="password"]/div[1]/div/div[1]/input', False, password) #Inputs password
                elementInteract('xpath', '//*[@id="passwordNext"]/div/button/span', True) #Clicks the next button
                sonicRun()
            except:
                elementInteract('xpath', '/html/body/div/form[1]/div/div/div[2]/div[1]/div/div/div/div/div[1]/div[3]/div/div/div/div[2]/div[2]/div/input[1]', False, username) #Inputs username
                elementInteract('xpath', '/html/body/div/form[1]/div/div/div[2]/div[1]/div/div/div/div/div[1]/div[3]/div/div/div/div[4]/div/div/div/div[2]/input', True) #Clicks next
                elementInteract('xpath', '/html/body/div/form[1]/div/div/div[2]/div[1]/div/div/div/div/div/div[3]/div/div[2]/div/div[3]/div/div[2]/input', False, password) #Inputs password
                elementInteract('xpath', '/html/body/div/form[1]/div/div/div[2]/div[1]/div/div/div/div/div/div[3]/div/div[2]/div/div[4]/div[2]/div/div/div/div/input', True) #Clicks next
                elementInteract('xpath', '/html/body/div/form/div/div/div[2]/div[1]/div/div/div/div/div/div[3]/div/div[2]/div/div[3]/div[2]/div/div/div[2]/input', True)
        except: 
            log('Failed to log in')   
            raise
        sleep(5)
        elemStatus = driver.find_element('xpath', '//*[@id="mCnz0"]/div/div/div[2]/span/div/div[3]/div/div/div/div[2]/ul/li[1]')
        driver.execute_script('arguments[0].click();', elemStatus) #Selects status
        if waitForExistance('//*[@id="mCnz0"]/div/div/div[2]/span/div/div[3]/div/div/div/div[2]/ul/li[1]/div/div[2]/div[2]/div/div/span/label[1]/div/div[3]/div', 10):
            elementInteract('xpath', '//*[@id="mCnz0"]/div/div/div[2]/span/div/div[3]/div/div/div/div[2]/ul/li[1]/div/div[2]/div[2]/div/div/span/label[1]/div/div[3]/div', True) #Selects All as a status filter
        else:return
        elementInteract('xpath', '//*[@id="mCnz0"]/div/div/div[2]/span/div/div[3]/div/div/div/div[2]/ul/li[1]/div/div[2]/div[3]/div/span/span', True) #Clicks apply to the status filter
        driver.implicitly_wait(0) #Disables implcitly_wait
        for i, asset in enumerate(assetChromebooks):
            elementInteract('xpath', '//*[@id="mCnz0"]/div/div/div[2]/span/div/div[3]/div/div/div/div[2]/ul/li[2]/div[2]/div[1]/input[2]', True) #clicks the blank space so the dropdown menu appear
            elemAssetID = driver.find_element('xpath', '/html/body/div[7]/c-wiz/div/div[1]/div/div/div[2]/div/div[2]/div[3]/div/div/div[2]/span/div/div[3]/div/div/div/div[2]/ul/li[2]/div[2]/div[2]/div/div/div/div[3]/div')
            driver.execute_script('arguments[0].click();', elemAssetID) #selects Asset ID as the search preference
            if waitForExistance('//*[@id="mCnz0"]/div/div/div[2]/span/div/div[3]/div/div/div/div[2]/ul/li[2]/div[2]/div[2]/div[2]/div/span/div/div/div/div/div[1]/div/div[1]/input', 10):
                elementInteract('xpath', '//*[@id="mCnz0"]/div/div/div[2]/span/div/div[3]/div/div/div/div[2]/ul/li[2]/div[2]/div[2]/div[2]/div/span/div/div/div/div/div[1]/div/div[1]/input', False, (str(asset), Keys.RETURN)) #searches the asset
            else:return
            try:
                sleep(1)
                elementInteract('xpath', '//*[@id="mCnz0"]/div/div/div[2]/span/div/div[3]/div/div/div/div[3]/div[2]/table/thead/tr/th[1]/span/div', True) #Clicks the checkbox
            except:
                print("Asset was not found")
                removeFilter()
                continue
            sleep(1)
            if not driver.find_element('xpath', '/html/body/div[7]/c-wiz/div/div[1]/div/div/div[2]/div/div[2]/div[3]/div/div/div[2]/span/div/div[3]/div/div/div/div[3]/div[2]/table/tbody/tr/td[5]/div[2]/span').get_attribute('innerHTML').__contains__('local'): #Checks if it's already in local OU
                elementInteract('xpath', '//*[@id="mCnz0"]/div/div/div[2]/span/div/div[3]/div/div/div/div[1]/div[2]/div[2]/div/div/div/div/span/span', True) #Clicks Move button to move it to a new OU
                sleep(1)
                elemLocal = driver.find_element('xpath', '/html/body/div[7]/div[5]/div/div[2]/span/div/span/span/div/div/c-wiz/div/div/c-wiz/div/div[3]/div[1]/ul/li/div/div/div[3]/div[2]')
                driver.execute_script("arguments[0].click();", elemLocal) #Clicks the Local OU
                sonicRun()
                sleep(0.1)
                elemMove = driver.find_element('xpath', '//*[@id="yDmH0d"]/div[5]/div/div[2]/span/div/div[2]/div[2]/span/span')
                driver.execute_script('arguments[0].click();', elemMove) #Moves it to the Local OU
                sleep(2)
                sonicRun()
                elementInteract('xpath', '//*[@id="mCnz0"]/div/div/div[2]/span/div/div[3]/div/div/div/div[3]/div[2]/table/thead/tr/th[1]/span/div', True) #Clicks checkbox
                sleep(0.5)
            else:
                print(f'{asset} is already in local')
                sonicRun()
            try:
                elemDisable = driver.find_element('xpath', '//*[@id="mCnz0"]/div/div/div[2]/span/div/div[3]/div/div/div/div[1]/div[2]/div[2]/div/div[1]/div[3]/div/span/span')
                driver.execute_script('arguments[0].click();', elemDisable) #Clicks Deprovision
                sleep(2)
                elemRetireFF = driver.find_element('xpath', '/html/body/div[7]/div[5]/div/div[2]/span/div/span/span/div[5]/span/table/tbody/tr[3]/td[1]/div/div[3]')
                driver.execute_script('arguments[0].click();', elemRetireFF) #selects Retiring From Fleet
                sleep(2)
                elemAccept = driver.find_element('xpath', '/html/body/div[7]/div[5]/div/div[2]/span/div/span/span/div[6]/div[1]/div[2]')
                driver.execute_script('arguments[0].click();', elemAccept) #Selects the checkbox saying you understand
                sleep(0.5)
                elemDeprovision = driver.find_element('xpath', '/html/body/div[7]/div[5]/div/div[2]/span/div/div[2]/div[2]/span')
                driver.execute_script('arguments[0].click();', elemDeprovision)
                sleep(2)
                print(f'{asset} has been deprovisioned')
                sonicRun()
            except Exception as error:
                print(f'{asset} is already deprovisioned')
                sonicRun()
            removeFilter()

    def submit_credentials():
        global username
        global password
        username = username_entry.get()
        password = password_entry.get()

        # Close the window
        window.destroy()
        
        #starts the main script
        mainGoogleAdmin()
    
    window = tk.Tk()

    # Creates a label and entry widget for the username
    username_label = tk.Label(window, text="Username:")
    username_label.pack()
    username_entry = tk.Entry(window)
    username_entry.pack()

    # Creates a label and entry widget for the password
    password_label = tk.Label(window, text="Password:")
    password_label.pack()
    password_entry = tk.Entry(window, show="*")
    password_entry.pack()

    # Creates a submit button
    submit_button = tk.Button(window, text="Submit", command=submit_credentials)
    submit_button.pack()

    # Runs the login info's event loop
    window.mainloop()

def iPads():
    #Finally deporovisioning
    def ipadFinal():
            try:
                driver.implicitly_wait(10)
                driver.switch_to.default_content()
                driver.switch_to.frame('MainPortal')
                for asset in assetIpads:
                    try:
                        elementInteract('xpath', '//*[@id="CWRoot"]/div/ui-main-pane/div[1]/div/div/div/div[2]/div/div/div/div/div/div[1]/div/div/div[1]/div/div/div/input', False, serialIpads[asset])
                        sleep(1)
                        driver.find_element(By.XPATH, '//*[@id="CWRoot"]/div/ui-main-pane/div[1]/div/div/div/div[2]/div/div/div/div/div/div[1]/div/div/div[3]/ul/li/div/div').click()
                        driver.find_element(By.XPATH, '//*[@id="CWRoot"]/div/ui-main-pane/div[1]/div/div/div/div[2]/div/div/div/div/div/div[2]/div/div/div[1]/div/button[2]/div').click()
                        driver.find_element(By.XPATH, '//*[@id="confirm-repair-mode"]').click()
                        driver.find_element(By.XPATH, '//*[@id="CWRoot"]/div/ui-pane/form/div/div/div/div[2]/div/span/div/button').click()
                        sleep(0.25)
                        driver.find_element(By.XPATH, '//*[@id="CWRoot"]/div/ui-main-pane/div[1]/div/div/div/div[2]/div/div/div/div/div/div[1]/div/div/div[1]/div/button').click()
                    except:
                        log(f'{asset} has already been released or an issue has prevented it from being released.')
                log('IPad deprovisioning has been completed')
                sonicRun()
            except:
                log('Something went wrong with searching the assets')
                sonicRun()
                raise
            return

    #Sheets to excel portion
    def ipadSTE():
        placeHolderTable = []
        corrector = 0
        if assetIpads:
            ipadGoogleSpreadsheet = serviceAccount.open(ipadSpreadsheetName)
            ipadInventorySheet = ipadGoogleSpreadsheet.worksheet('iPad Inv')
            ipadAllInfo = ipadInventorySheet.get_all_records()
            for asset in assetIpads:
                placeHolderTable.append(asset)
            for i, rows in enumerate(ipadAllInfo):
                if rows['Asset'] in placeHolderTable:
                    sleep(0.9)
                    placeHolderTable.remove(rows['Asset'])
                    sellWorksheet.append(list(rows.values())) #Enters the data into the next open row in the excel sheet
                    sellWorkbook.save(workbookSell) #Saves the workbook
                    log( "Found and Written " + str(rows['Asset']) + '\n') #Enters that it found the certain asset tag
                    sonicRun()
                    serialIpads[list(rows.values())[0]] = list(rows.values())[1]
                    ipadInventorySheet.delete_rows(i + 2 - corrector)
                    corrector += 1
            for leftOvers in placeHolderTable: #Every asset that wasn't found
                log('Did not find asset# '+str(leftOvers) + ' in inventory' + '\n') #Writes the asset into logs
                sonicRun()
        else:
            log('No iPad targets')
            sonicRun()
            exit()
        ipadFinal()
        return
        
    def ipadMid():
                buttons = driver.find_elements(By.TAG_NAME, 'button')
                for button in buttons:
                    try:
                        if button.text == 'Not Now':
                            button.click()
                        else:
                            continue
                    except:
                        print('Garbage collection is working')
                ipadSTE()
                return
    def ipadMain():
        global driver
        global serialIpads
        placeHolderTable = []
        corrector = 0
        options = Options()
        options.add_argument('--incgonito')
        options.add_experimental_option('detach', True) #Closes window on completion
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()),
        options = options) #Chooses Chrome as the browser
        driver.implicitly_wait(3) #Waits 10 seconds to find an element befor sending an error
        driver.get("https://school.apple.com/#/main/devices")
        try:
            driver.switch_to.frame('aid-auth-widget')
            driver.find_element(By.XPATH, '//*[@id="account_name_text_field"]').send_keys(appleID)
            driver.find_element(By.XPATH, '//*[@id="account_name_text_field"]').send_keys(Keys.ENTER)
            driver.find_element(By.XPATH, '//*[@id="password_text_field"]').send_keys(applePassword)
            driver.find_element(By.XPATH, '//*[@id="password_text_field"]').send_keys(Keys.ENTER)
            phoneNumberText = driver.find_element(By.XPATH, '//*[@id="stepEl"]/hsa2/div/verify-phone/div/div/div[2]/p/span').text
            try:
                def submit_phoneCode():
                    global phoneNumber
                    phoneNumber = phoneCode_entry.get()
                    # Close the window
                    phoneCode.destroy()
                    driver.find_element(By.XPATH, '//*[@id="char0"]').send_keys(phoneNumber)
                    sonicRun()
                    ipadMid()
                    root.quit()
                    sonicRun()
                    return
                    
                phoneCode = tk.Tk()

                # Creates a label and entry widget for the username
                phoneCode_label = tk.Label(phoneCode, text="Code sent to " + str(phoneNumberText))
                phoneCode_label.pack()
                phoneCode_entry = tk.Entry(phoneCode)
                phoneCode_entry.pack()

                # Creates a submit button
                submit_button = tk.Button(phoneCode, text="Submit", command=submit_phoneCode)
                submit_button.pack()
                # Runs the login info's event loop
                phoneCode.mainloop()
            except:
                log('an issue occured with the phone number entry')
                quit()

        except:
            log('login issue occured')
            quit()

    def submit_credentialsASM():
        global appleID
        global applePassword
        appleID = username_entry.get()
        applePassword = password_entry.get()

        # Close the window
        window.destroy()
        ipadMain()
        sonicRun()
        root.quit()
        return

    window = tk.Tk()

    # Creates a label and entry widget for the Apple ID
    username_label = tk.Label(window, text="Apple ID:")
    username_label.pack()
    username_entry = tk.Entry(window)
    username_entry.pack()

    # Creates a label and entry widget for the password
    password_label = tk.Label(window, text="Password:")
    password_label.pack()
    password_entry = tk.Entry(window, show="*")
    password_entry.pack()

    # Creates a submit button
    submit_button = tk.Button(window, text="Submit", command=submit_credentialsASM)
    submit_button.pack()

    # Runs the login info's event loop
    window.mainloop()

#The Main Gui is beyond this point
def run():
    global currentlyRunning
    global assetChromebooks
    global rowsFound
    global tagsNotFound
    global togs
    if not currentlyRunning:
        log('Started new assignment at ' + str(datetime.now()))
        currentlyRunning = True
        runButton['background'] = 'Yellow'
        try:
            openSellWorkbook()
            getTargets()
        except:
            log('Somebody has the Sell or Targets Workbook open')
        try:
            root.update()
            global percentDone
            percentDone = 0
            for func in togs:
                if func == sheetstoexcel:
                    func(invSheetName.get())
                else:
                    func()
                percentDone += 1/len(togs)*100
                ringText.configure(text=str(percentDone) + '%')
                root.update()
            runButton['background'] = 'LightGreen'
            rowsFound = [] #Table for the assets found and their information
            tagsNotFound = []
            sonicLabel.configure(image=sonicFinishedPhoto)
            rankLabel.configure(image=rankSPhoto)
            rankLabel.place(x = 40, y= 110)
            
        except Exception as error:
            log('An error occured. Run script in VS and see the error.')
            print(error)
            sonicLabel.configure(image=sonicStopPhoto)
            rankLabel.configure(image=rankFPhoto)
            rankLabel.place(x = 15, y = 110)
        finally:
            log('Finished assignment at ' + str(datetime.now()))
            togs = []
            destinyButton['background'] = 'Tomato'
            invToSellButton['background'] = 'Tomato'
            forceRetireButton['background'] = 'Tomato'
            googleAdminButton['background'] = 'Tomato'

        currentlyRunning = False
        print('Finished')
    else:
         log('Something went wrong, restart instance.')

def warningPermanent():
    yesno = messagebox.askyesno('Confirmation', "Warning! This is a permanent! Do you wish to proceed?")
    if yesno == True:
        return True
    else:
        return False

def destinyB():
    if destinyButton['background'] == 'Tomato':
        togs.append(destinyFull)
        destinyButton['background'] = 'LightGreen'
    else:
        togs.remove(destinyFull)
        destinyButton['background'] = 'Tomato'

def invToSellB():
    if invToSellButton['background'] == 'Tomato':
        togs.append(sheetstoexcel)
        invToSellButton['background'] = 'LightGreen'
    else:
        togs.remove(sheetstoexcel)
        invToSellButton['background'] = 'Tomato'

def ipadB():
    if ipadButton['background'] == 'Tomato':
        if warningPermanent():
            togs.append(iPads)
            ipadButton['background'] = 'LightGreen'
        else:
            return
    else:
        togs.remove(iPads)
        ipadButton['background'] = 'Tomato'

def forceRetireTog():
    global forceRetire
    if forceRetireButton['background'] == 'Tomato':
        forceRetire = True
        forceRetireButton['background'] = 'LightGreen'
    else:
        forceRetireButton['background'] = 'Tomato'
        forceRetire = False

def googleAdminB():
    if googleAdminButton['background'] == 'Tomato':
        if warningPermanent():
            togs.append(googleAdmin)
            googleAdminButton['background'] = 'LightGreen'
        else:
            return
    else:
        togs.remove(googleAdmin)
        googleAdminButton['background'] = 'Tomato'

root = tk.Tk()
root.geometry('750x600')
root.title("Owen's Auto Recycler")

sheetSelectionText = tk.Label(root, text = 'Select possible location of assets', font = ('Arial', 13, 'bold'))
sheetSelectionText.place(x=235, y=2)

buttonsInfoText = tk.Label(root, text = 'Select all of the functions that are needed. Any combination of buttons is functional.',
 font = ('Arial', 13, 'bold'))
buttonsInfoText.place(x=60, y=80)

buttonframeCol2 = tk.Frame(root)
buttonframeCol2.columnconfigure(0, weight=1)
buttonframeCol2.columnconfigure(1, weight=1)
buttonframeCol2.columnconfigure(2, weight=1)
buttonframeCol2.place(x=260, y=110)

buttonframeCol3 = tk.Frame(root)
buttonframeCol3.columnconfigure(0, weight=2)
buttonframeCol3.columnconfigure(1, weight=2)
buttonframeCol3.place(x=510, y=110)

destinyButton = tk.Button(buttonframeCol2, text = 'Destiny check', font=('Arial', 14, 'bold'), command=destinyB,
 foreground='Black', background='Tomato', height=1, width=18)
destinyButton.grid(row=0, column=0, sticky=tk.W+tk.E, pady=2)

invToSellButton = tk.Button(buttonframeCol2, text = 'Inventory To Sell', font=('Arial', 14, 'bold'),
 command=invToSellB, foreground = 'Black', background='Tomato', height=1, width=18)
invToSellButton.grid(row=3, column=0, pady=2)

ipadButton = tk.Button(buttonframeCol3, text = 'Full iPad Deprovision', font=('Arial', 14, 'bold'),
 command=ipadB, foreground = 'Black', background='Tomato', height=1, width=18)
ipadButton.grid(row=1, column=0, pady=2)

runButton = tk.Button(root, text = 'Start', font=('Arial', 14, 'bold'), background = 'LightGreen',
 command=run)
runButton.place(x=340, y=240)

forceRetireButton = tk.Button(buttonframeCol2, text = 'Force Retire', font=('Arial', 14, 'bold'),
 command=forceRetireTog, foreground = 'Black', background='Tomato', height=1, width=18)
forceRetireButton.grid(row=1, column=0, pady=2)

googleAdminButton = tk.Button(buttonframeCol3, text = 'Google Admin', font=('Arial', 14, 'bold'),
 command=googleAdminB, foreground = 'Black', background='Tomato', height=1, width=18)
googleAdminButton.grid(row = 0, column = 0, pady=2)

textLogs = ScrolledText(root, width=90, height= 17)
textLogs.place(x = 10, y = 310)

invSheetName = StringVar()
invSheetName.set('No Sheet Selected')
allSheets = chromeGoogleSpreadsheet.worksheets()
sheetChoices = []
exclusions = ['Main Numbers', 'EOL', 'Parts', 'Missing', 'Destiny Imports', 'School Matrix',
 'Student Enrollment', 'DRC Device ID', 'Combo Lists', 'Network Color Scheme', 'DICS']
for sheet in allSheets:
    if sheet.title not in exclusions:
        sheetChoices.append(sheet.title)
def onOptionSelected(event):
    invSheetDd.configure(background='LightGreen')
invSheetDd = OptionMenu(root, invSheetName, *sheetChoices, command=onOptionSelected)
invSheetDd.place(x = 250, y = 25)
invSheetDd.configure(background='Tomato', font=('Arial', 14, 'bold'), height=1, width=18)

#Loads Sonic sprites
sonicIdlePath = Image.open()
sonicIdlePhoto = ImageTk.PhotoImage(sonicIdlePath)

sonicRunPath_1 = Image.open()
sonicRunPhoto_1 = ImageTk.PhotoImage(sonicRunPath_1)

sonicRunPath_2 = Image.open()
sonicRunPhoto_2 = ImageTk.PhotoImage(sonicRunPath_2)

sonicRunPath_3 = Image.open()
sonicRunPhoto_3 = ImageTk.PhotoImage(sonicRunPath_3)

sonicRunPath_4 = Image.open()
sonicRunPhoto_4 = ImageTk.PhotoImage(sonicRunPath_4)

sonicRunPath_5 = Image.open()
sonicRunPhoto_5 = ImageTk.PhotoImage(sonicRunPath_5)

sonicRunPath_6 = Image.open()
sonicRunPhoto_6 = ImageTk.PhotoImage(sonicRunPath_6)

sonicRunPath_7 = Image.open()
sonicRunPhoto_7 = ImageTk.PhotoImage(sonicRunPath_7)

sonicRunPath_8 = Image.open()
sonicRunPhoto_8 = ImageTk.PhotoImage(sonicRunPath_8)

sonicStopPath = Image.open()
sonicStopPhoto = ImageTk.PhotoImage(sonicStopPath)

sonicFinishedPath = Image.open()
sonicFinishedPhoto = ImageTk.PhotoImage(sonicFinishedPath)

sonicRingPath = Image.open()
sonicRingPhoto = ImageTk.PhotoImage(sonicRingPath)

rankSPath = Image.open()
rankSPhoto = ImageTk.PhotoImage(rankSPath)

rankFPath = Image.open()
rankFPhoto = ImageTk.PhotoImage(rankFPath)

sonicRunFull = [sonicRunPhoto_1, sonicRunPhoto_2, sonicRunPhoto_3, sonicRunPhoto_4, sonicRunPhoto_5,
sonicRunPhoto_6, sonicRunPhoto_7, sonicRunPhoto_8]

#Puts sonic on the gui
sonicLabel = tk.Label(root, image=sonicIdlePhoto)
sonicLabel.place(x = 75, y = 100)

#Ring counter (percentage done)
ringLabel = tk.Label(root, image = sonicRingPhoto)
ringLabel.place(x = 70 ,y = 260)
ringText = tk.Label(root, text = str(percentDone) + '%', font=('papyrus', 14, 'bold'), foreground='darkgoldenrod')
ringText.place (x = 110, y = 257)

#Rank on the left side after completion
rankLabel = tk.Label(root, image=None)

root.mainloop()
