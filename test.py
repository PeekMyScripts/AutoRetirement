from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from selenium.webdriver.common.keys import Keys
import gspread
import tkinter as tk
from tkinter import *
from tkinter.scrolledtext import *
from tkinter.filedialog import askopenfilename
from datetime import datetime
from time import sleep

pathServiceAccount = r"C:\Users\tech3\Documents\AutoRetirement\mass-device-automation-tool-ac1772a2fd3e.json" #1
pathTxt = r"C:\Users\tech3\Documents\AutoRetirement\logs.txt" #2
spreadsheetName = "Copy of Copy of Inv Tracker Copy 2022-12-23 07:47:20" #3
sellSheetName = 'Current Year'
workbookSell = r"C:\Users\tech3\Documents\AutoRetirement\Sell.xlsx" #5
workbookTargetAssets = r"C:\Users\tech3\Documents\AutoRetirement\assetTags.xlsx" #6
currentlyRunning = False #determines if it's currently running or not
togs = [] #Table for all of the active toggles
serviceAccount = gspread.service_account(filename = pathServiceAccount) #loads the service account
googleSpreadsheet = serviceAccount.open(spreadsheetName) #Google spreadsheet's name
targetWorkbook = load_workbook(workbookTargetAssets) #Loads excel worksheet file for 
targetWorksheet = targetWorkbook.active #Gets active worksheet to pull asset tags from
sellWorkbook = load_workbook(workbookSell) #Loads sell excel file
sellWorksheet = sellWorkbook[sellSheetName] #Active worksheet inside of the workbook
sheetName = 'Chromebooks'
forceRetire = False #Forcefully retires device, defaults to false
invSheetName = None
sheets = googleSpreadsheet.worksheet('Chromebooks')
sheetsToLook = [googleSpreadsheet.worksheet('Chromebooks')]
assets = [] #Target asset tags to find
rowsFound = [] #Table for the assets found and their information
tagsNotFound = [] #Table for assets not found
allInfo = None
driver = []
def elementInteract(method, element, clickOrKeys, keysToSend = None):
    if clickOrKeys:
        driver.find_element(method, element).click()
    else:
        driver.find_element(method, element).send_keys(keysToSend)
def waitForExistance(elementXPath, waittime):
    try:
        waitFor = WebDriverWait(driver, waittime).until(EC.presence_of_element_located((By.XPATH, elementXPath)))
        return True
    except:
        return False
def waitUntilClickable(element, waittime):
    try:
        WebDriverWait(driver, waittime).until(EC.element_to_be_clickable((By.XPATH, element)))
        return True
    except:
        return False
def openSellWorkbook():
    global sellWorkbook
    global sellWorksheet
    sellWorkbook = load_workbook(workbookSell)
    sellWorksheet = sellWorkbook[sellSheetName]
def getAllRec(sheet):
    return sheet.get_all_records()
for index, cell in enumerate(targetWorksheet['A']): #Grabs all the values inside of the recieving worksheet
    if cell.value != None: #Checks for empty cells
        assets.insert(index, cell.value) #Inserts the values into "assets"
def sheetstoexcel(sheetname): #full function for moving information from the inventory tracker to the sell file
    global assets
    global rowsFound
    global tagsNotFound
    inventorySheet = googleSpreadsheet.worksheet(sheetname) #Sheet open inside of the main spreadsheet
    allInfo = getAllRec(inventorySheet)
    placeHolderTable = []
    corrector = 0
    for asset in assets:
        placeHolderTable.append(asset)
    for i, rows in enumerate(allInfo):
        if rows['Asset'] in placeHolderTable:
            sleep(0.9)
            placeHolderTable.remove(rows['Asset'])
            sellWorksheet.append(list(rows.values())) #Enters the data into the next open row in the excel sheet
            sellWorkbook.save(workbookSell) #Saves the workbook
            log( "Found and Written " + str(rows['Asset']) + '\n') #Enters that it found the certain asset tag
            inventorySheet.delete_rows(i + 2 - corrector)
            corrector += 1
            print(corrector)
    for leftOvers in placeHolderTable: #Every asset that wasn't found
        log('Did not find asset# '+str(leftOvers) + ' in inventory' + '\n') #Writes the asset into logs

def log(text):
    txtWrite = open(pathTxt, 'a')
    txtWrite.write(text + '\n') #Writes in logs
    textLogs.insert(END, text + '\n')
    txtWrite.close()

def removeFilter():
    elemRemoveFilter = driver.find_element('xpath', '//*[@id="mCnz0"]/div/div/div[2]/span/div/div[3]/div/div/div/div[2]/ul/li[2]/div')
    driver.execute_script('arguments[0].click();', elemRemoveFilter) #Removes the asset tag searched

def destinyFull():
    global forceRetire
    options = Options()

    options.add_experimental_option('detach', True) #closes window on completion
    options.headless = False #If selected true, it runs in the background

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()),
    options = options) #Chooses Chrome as the browser

    driver.get("https://library.seymour.k12.wi.us/district/servlet/presentdistrictloginform.do") #Goes to the destiny login page

    driver.find_element('xpath', '//*[@id="Login"]').click() #Clicks to go to the login screen
    driver.find_element('xpath', '//*[@id="ID_loginName"]').send_keys('destinyadmin') #Enters username
    driver.find_element('xpath', '//*[@id="ID_password"]').send_keys('destiny@dmin') #Enters password
    driver.find_element('xpath', '/html/body/table/tbody/tr[4]/td/table/tbody/tr/td/form/table/tbody/tr[3]/td/table/tbody/tr/td/table/tbody/tr[4]/td/input[1]').click() #Clicks login
    driver.find_element('xpath', '//*[@id="rightTableSites_0"]/tbody/tr/td/ul/li/span/a[1]').click() #Clicks SMS as the location
    driver.get('https://library.seymour.k12.wi.us/cataloging/servlet/presentadvancedsearchredirectorform.do?l2m=Library%20Search&tm=TopLevelCatalog') #Goes to the catalog search
    driver.find_element('xpath', '//*[@id="Resource Search"]').click() #Clicks on the resource search
    driver.find_element('xpath', '//*[@id="searchFieldsTable"]/tbody/tr/td[1]/select/optgroup[2]/option[1]').click() #Chooses the barcode option
    def checkExistance(xpath): #Checks to see if a element exists
        try:
            driver.find_element('xpath', xpath)
        except NoSuchElementException:
                return False
        return True
    for i, asset in enumerate(assets): #Goes through this process with each one of the assets
        if len(str(asset)) == 4:
            driver.find_element('xpath', '/html/body/table[2]/tbody/tr[3]/td/table/tbody/tr/td[2]/table/tbody/tr/td[2]/table/tbody/tr/td/form/table[2]/tbody/tr/td/table/tbody/tr[2]/td/table/tbody/tr[2]/td[2]/table/tbody/tr/td[2]/input').send_keys('0' + str(asset)) #Types in the asset# + a 0 at the start
        elif len(str(asset)) > 4:
            driver.find_element('xpath', '/html/body/table[2]/tbody/tr[3]/td/table/tbody/tr/td[2]/table/tbody/tr/td[2]/table/tbody/tr/td/form/table[2]/tbody/tr/td/table/tbody/tr[2]/td/table/tbody/tr[2]/td[2]/table/tbody/tr/td[2]/input').send_keys(str(asset))
        driver.find_element('xpath', '//*[@id="buttonSearch"]').click() #Searches
        if checkExistance('//*[@id="messageBox"]/tbody/tr[1]/td[2]'): #If the chromebook isn't already at the SMS site
            if driver.find_element('xpath', '//*[@id="messageBox"]/tbody/tr[1]/td[2]').get_attribute('innerHTML').__contains__('Please note'): #Makes sure that it's asking the question if you wanted to move it
                driver.find_element('xpath', '//*[@id="divToHide"]/input[1]').click() #Moves it to the SMS site
        if checkExistance('//*[@id="messageBox"]/tbody/tr/td'): #Checks to see if the element saying the chromebook isn't inside of destiny exists
            log(str(asset) + ' is not avalaible') #Writes in logs saying the Chromebook wasn't avalaible in destiny
            driver.find_element('xpath', '//*[@id="Resource Search"]').click() #Clicks back to the resource search
            continue #Restarts the process with the next asset in line
        if driver.find_element('xpath', '//*[@id="itemTable_0"]/tbody/tr[2]/td[3]/a').get_attribute('innerHTML').__contains__('Available'): #Checks to see if the Chromebook isn't checked out or already retired
            driver.find_element('xpath', '//*[@id="itemTable_0"]/tbody/tr[2]/td[6]/a[2]/img').click() #clicks on the edit button
            log(str(asset) + ' Has been retired ' + ':' + ')' + '\n')
        elif driver.find_element('xpath', '//*[@id="itemTable_0"]/tbody/tr[2]/td[3]/a').get_attribute('innerHTML').__contains__('Checked') and forceRetire:
            driver.find_element('xpath', '//*[@id="TopLevelCirculation"]').click()
            driver.find_element('xpath', '//*[@id="Check In Items"]').click()
            if len(str(asset)) == 4:
                driver.find_element('xpath', '//*[@id="headerTable"]/tbody/tr/td[2]/input').send_keys('0' + str(asset))
            else:
                driver.find_element('xpath', '//*[@id="headerTable"]/tbody/tr/td[2]/input').send_keys(str(asset))
            driver.find_element('xpath', '//*[@id="go"]').click()
            if checkExistance('//*[@id="messageBox"]/tbody/tr/td[2]'):
                driver.find_element('xpath', '//*[@id="ComponentsTable"]/tbody/tr[5]/td/input[1]').click()
            log(str(asset) + ' was  ' + driver.find_element('xpath', '/html/body/table[2]/tbody/tr[3]/td/table/tbody/tr/td[2]/table/tbody/tr/td[2]/table/tbody/tr/td/form/div[3]/table/tbody/tr[3]/td[2]/span[1]').get_attribute('innerHTML') + ' but is now checked in '+ '\n')
            driver.find_element('xpath', '//*[@id="TopLevelCatalog"]').click()
            driver.find_element('xpath', '//*[@id="Resource Search"]').click()
            driver.find_element('xpath', '//*[@id="searchFieldsTable"]/tbody/tr/td[1]/select/optgroup[2]/option[1]').click()
            if len(str(asset)) == 4:
                print(len(str(asset)))
                driver.find_element('xpath', '//*[@id="searchFieldsTable"]/tbody/tr/td[2]/input').send_keys('0' + str(asset))
            else:
                driver.find_element('xpath', '//*[@id="searchFieldsTable"]/tbody/tr/td[2]/input').send_keys(str(asset))
            driver.find_element('xpath', '//*[@id="buttonSearch"]').click()
            driver.find_element('xpath', '//*[@id="itemTable_0"]/tbody/tr[2]/td[6]/a[2]/img').click()
        else: #If it isn't available
            log(str(asset) + ' is marked as ' + str(driver.find_element('xpath', '//*[@id="itemTable_0"]/tbody/tr[2]/td[3]/a').get_attribute('innerHTML')) + ' and needs to be done manually') #Writes in logs
            driver.find_element('xpath', '//*[@id="Resource Search"]').click() #Clicks back on reasource search
            continue
        driver.find_element('xpath', '//*[@id="tableMain"]/tbody/tr[4]/td[2]/table/tbody/tr/td[1]/select/option[9]').click() #Checks status to retired
        driver.find_element('xpath', '//*[@id="saveCopy"]').click() #Clicks save
        driver.find_element('xpath', '//*[@id="Resource Search"]').click() #Goes back to resource search
        driver.find_element('xpath', '//*[@id="searchFieldsTable"]/tbody/tr/td[1]/select/optgroup[2]/option[1]').click() #clicks on barcode option

def googleAdmin():
    global driver
    options = Options()

    options.add_experimental_option('detach', True) #Closes window on completion
    options.headless = False #If selected true, it runs in the background

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()),
    options = options) #Chooses Chrome as the browser

    driver.get("https://admin.google.com/ac/chrome/devices/")
    elementInteract('xpath', '//*[@id="identifierId"]', False, 'tech3@seymour.k12.wi.us')
    elementInteract('xpath', '//*[@id="identifierNext"]/div/button/span', True)
    if waitForExistance('//*[@id="view_container"]/div/div/div[2]/div/div[2]/div/div[1]/div/div/button/span', 20):
        elementInteract('xpath', '//*[@id="view_container"]/div/div/div[2]/div/div[2]/div/div[1]/div/div/button/span', True)
    else:
        log('Something went wrong with login')
        quit()
    sleep(5)
    elemStatus = driver.find_element('xpath', '//*[@id="mCnz0"]/div/div/div[2]/span/div/div[3]/div/div/div/div[2]/ul/li[1]')
    driver.execute_script('arguments[0].click();', elemStatus) #Selects status
    if waitForExistance('//*[@id="mCnz0"]/div/div/div[2]/span/div/div[3]/div/div/div/div[2]/ul/li[1]/div/div[2]/div[2]/div/div/span/label[1]/div/div[3]/div', 10):
        elementInteract('xpath', '//*[@id="mCnz0"]/div/div/div[2]/span/div/div[3]/div/div/div/div[2]/ul/li[1]/div/div[2]/div[2]/div/div/span/label[1]/div/div[3]/div', True) #Selects All as a status filter
    else:return
    elementInteract('xpath', '//*[@id="mCnz0"]/div/div/div[2]/span/div/div[3]/div/div/div/div[2]/ul/li[1]/div/div[2]/div[3]/div/span/span', True) #Clicks apply to the status filter
    for i, asset in enumerate(assets):
        sleep(1)
        elementInteract('xpath', '//*[@id="mCnz0"]/div/div/div[2]/span/div/div[3]/div/div/div/div[2]/ul/li[2]/div[2]/div[1]/input[2]', True) #clicks the blank space so the dropdown menu appears
        sleep(0.5)
        elemAssetID = driver.find_element('xpath', '/html/body/div[7]/c-wiz/div/div[1]/div/div/div[2]/div/div[2]/div[3]/div/div/div[2]/span/div/div[3]/div/div/div/div[2]/ul/li[2]/div[2]/div[2]/div/div/div/div[3]/div')
        driver.execute_script('arguments[0].click();', elemAssetID) #selects Asset ID as the search preference
        if waitForExistance('//*[@id="mCnz0"]/div/div/div[2]/span/div/div[3]/div/div/div/div[2]/ul/li[2]/div[2]/div[2]/div[2]/div/span/div/div/div/div/div[1]/div/div[1]/input', 10):
            elementInteract('xpath', '//*[@id="mCnz0"]/div/div/div[2]/span/div/div[3]/div/div/div/div[2]/ul/li[2]/div[2]/div[2]/div[2]/div/span/div/div/div/div/div[1]/div/div[1]/input', False, (str(asset), Keys.RETURN)) #searches the asset
        else:return
        try:
            sleep(1)
            elementInteract('xpath', '//*[@id="mCnz0"]/div/div/div[2]/span/div/div[3]/div/div/div/div[3]/div[2]/table/thead/tr/th[1]/span/div', True) #Clicks the checkbox
        except:
            log(f"{asset} was not found by searching in google admin")
            removeFilter()
            continue
        sleep(1)
        if not driver.find_element('xpath', '/html/body/div[7]/c-wiz/div/div[1]/div/div/div[2]/div/div[2]/div[3]/div/div/div[2]/span/div/div[3]/div/div/div/div[3]/div[2]/table/tbody/tr/td[5]/div[2]/span').get_attribute('innerHTML').__contains__('local'): #Checks if it's already in local OU
            elementInteract('xpath', '//*[@id="mCnz0"]/div/div/div[2]/span/div/div[3]/div/div/div/div[1]/div[2]/div[2]/div/div/div/div/span/span', True) #Clicks Move button to move it to a new OU
            sleep(1)
            elemLocal = driver.find_element('xpath', '/html/body/div[7]/div[5]/div/div[2]/span/div/span/span/div/div/c-wiz/div/div/c-wiz/div/div[3]/div[1]/ul/li/div/div/div[3]/div[2]')
            driver.execute_script("arguments[0].click();", elemLocal) #Clicks the Local OU
            sleep(0.1)
            elemMove = driver.find_element('xpath', '//*[@id="yDmH0d"]/div[5]/div/div[2]/span/div/div[2]/div[2]/span/span')
            driver.execute_script('arguments[0].click();', elemMove) #Moves it to the Local OU
            sleep(2)
            elementInteract('xpath', '//*[@id="mCnz0"]/div/div/div[2]/span/div/div[3]/div/div/div/div[3]/div[2]/table/thead/tr/th[1]/span/div', True) #Clicks checkbox
            sleep(0.5)
            log(f'{asset} has been put in local')
        else:
            log(f'{asset} is already in local')
        try:
            elemDisable = driver.find_element('xpath', '//*[@id="mCnz0"]/div/div/div[2]/span/div/div[3]/div/div/div/div[1]/div[2]/div[2]/div/div[1]/div[3]/div/span/span')
            driver.execute_script('arguments[0].click();', elemDisable) #Clicks Deprovision
            sleep(2)
            elemRetireFF = driver.find_element('xpath', '/html/body/div[7]/div[5]/div/div[2]/span/div/span/span/div[5]/span/table/tbody/tr[3]/td[1]/div/div[3]')
            driver.execute_script('arguments[0].click();', elemRetireFF) #selects Retiring From Fleet
            sleep(2)
            elemAccept = driver.find_element('xpath', '/html/body/div[7]/div[5]/div/div[2]/span/div/span/span/div[6]/div[1]/div[2]')
            driver.execute_script('arguments[0].click();', elemAccept) #Selects the checkbox saying you understand
            sleep(0.5)
            elemDeprovision = driver.find_element('xpath', '/html/body/div[7]/div[5]/div/div[2]/span/div/div[2]/div[2]/span')
            driver.execute_script('arguments[0].click();', elemDeprovision)
            sleep(2)
            log(f'{asset} has been deprovisioned')
        except Exception as error:
            log(f'{asset} is already deprovisioned')
        removeFilter()

#The Gui is beyond this point
def run():
    global currentlyRunning
    global assets
    global rowsFound
    global tagsNotFound
    global togs
    if not currentlyRunning:
        log('Started new assignment at ' + str(datetime.now()))
        currentlyRunning = True
        runButton['background'] = 'Yellow'
        sleep(1)
        try:
            openSellWorkbook()
        except:
            log('Somebody has the Sell Workbook open')
        try:
            for func in togs:
                if func == sheetstoexcel:
                    func(invSheetName.get())
                else:
                    func()
            runButton['background'] = 'Green'
            assets = [] #Target asset tags to find
            rowsFound = [] #Table for the assets found and their information
            tagsNotFound = []
            
        except Exception as error:
            log('An error occured. Run script in VS and see the error.')
            print(error)
        finally:
            log('Finished assignment at ' + str(datetime.now()))
            togs = []
            destinyButton['background'] = 'Red'
            invToSellButton['background'] = 'Red'
            forceRetireButton['background'] = 'Red'
            googleAdminButton['background'] = 'Red'

        currentlyRunning = False
        print('Finished')
    else:
         log('Something went wrong, restart instance.')


def destinyB():
    if destinyButton['background'] == 'Red':
        togs.append(destinyFull)
        destinyButton['background'] = 'Green'
    else:
        togs.remove(destinyFull)
        destinyButton['background'] = 'Red'
def invToSellB():
    if invToSellButton['background'] == 'Red':
        togs.append(sheetstoexcel)
        invToSellButton['background'] = 'Green'
    else:
        togs.remove(sheetstoexcel)
        invToSellButton['background'] = 'Red'
def forceRetireTog():
    global forceRetire
    if forceRetireButton['background'] == 'Red':
        forceRetire = True
        forceRetireButton['background'] = 'Green'
    else:
        forceRetireButton['background'] = 'Red'
        forceRetire = False

def googleAdminB():
    if googleAdminButton['background'] == 'Red':
        togs.append(googleAdmin)
        googleAdminButton['background'] = 'Green'
    else:
        togs.remove(googleAdmin)
        googleAdminButton['background'] = 'Red'
root = tk.Tk()
root.geometry('1200x700')
root.title("Owen's Auto Recycler")
root.iconbitmap(r"C:\Users\tech3\Documents\AutoRetirement\icon.ico")

buttonframeCol1 = tk.Frame(root)
buttonframeCol1.columnconfigure(0, weight=2)
buttonframeCol1.columnconfigure(1, weight=2)
buttonframeCol1.place(x=10, y=60)

buttonframeCol2 = tk.Frame(root)
buttonframeCol2.columnconfigure(0, weight=2)
buttonframeCol2.columnconfigure(1, weight=2)
buttonframeCol2.place(x=260, y=60)

buttonframeCol3 = tk.Frame(root)
buttonframeCol3.columnconfigure(0, weight=2)
buttonframeCol3.columnconfigure(1, weight=2)
buttonframeCol3.place(x=460, y=60)

destinyButton = tk.Button(buttonframeCol2, text = 'Destiny check', font=('Arial', 18), command=destinyB, foreground='DarkBlue', background='Red')
destinyButton.grid(row=0, column=0, sticky=tk.W+tk.E)

invToSellButton = tk.Button(buttonframeCol1, text = 'Inventory To Sell', font=('Arial', 18), command=invToSellB, foreground = 'DarkBlue', background='Red')
invToSellButton.grid(row=0, column=0)

runButton = tk.Button(root, text = 'Start', font=('Arial', 18), background = 'Green', command=run)
runButton.place(x=310, y=10)

forceRetireButton = tk.Button(buttonframeCol2, text = 'Force Retire', font=('Arial', 18), command=forceRetireTog, foreground = 'DarkBlue', background='Red')
forceRetireButton.grid(row=1, column=0)

googleAdminButton = tk.Button(buttonframeCol3, text = 'Google Admin', font=('Arial', 18), command=googleAdminB, foreground = 'DarkBlue', background='Red')
googleAdminButton.grid(row = 2, column = 0)
textLogs = ScrolledText(root, width=60, height= 40)
textLogs.place(x = 700, y = 10)

invSheetName = StringVar()
invSheetName.set('Choose sheet to look in')
allSheets = googleSpreadsheet.worksheets()
sheetChoices = []
for sheet in allSheets:
    sheetChoices.append(sheet.title)

invSheetDd = OptionMenu(root, invSheetName, *sheetChoices)
invSheetDd.place(x = 20, y = 10)
root.mainloop()